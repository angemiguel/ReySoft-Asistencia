import csv
from io import BytesIO, StringIO
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func, select, update
from sqlalchemy.orm import Session

from app.core.permissions import ensure_school_admin, ensure_school_user
from app.database.session import get_db
from app.dependencies.auth import get_current_user
from app.models import Course, Guardian, Student, StudentGuardian, User
from app.schemas.student import (
    StudentCreate,
    StudentGuardianCreate,
    StudentGuardianResponse,
    StudentImportResponse,
    StudentResponse,
    StudentUpdate,
)
from app.services.audit import create_audit_log
from app.utils.phone import clean_phone_number

router = APIRouter(prefix="/students", tags=["students"])

STUDENT_EXCEL_HEADERS = [
    "nombre_completo",
    "codigo",
    "curso",
    "seccion",
    "anio_academico",
    "activo",
    "tutor_principal",
    "tutor_principal_telefono",
]
STUDENT_IMPORT_REQUIRED_HEADERS = [
    "nombre_completo",
    "codigo",
    "curso",
    "seccion",
    "anio_academico",
    "activo",
    "tutor_principal_telefono",
]
EXCEL_HEADER_FILL = "FF16A34A"
EXCEL_HEADER_FONT = "FFFFFF"
EXCEL_MAX_COLUMN_WIDTH = 60
CSV_BOM = "\ufeff"


def _get_student_or_404(db: Session, student_id: UUID, organization_id: UUID) -> Student:
    student = db.scalar(select(Student).where(Student.id == student_id, Student.organization_id == organization_id))
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Estudiante no encontrado.")
    return student


def _ensure_course_in_organization(db: Session, course_id: UUID, organization_id: UUID) -> Course:
    course = db.scalar(select(Course).where(Course.id == course_id, Course.organization_id == organization_id))
    if not course:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El curso no pertenece a tu centro educativo.")
    return course


def _ensure_guardians_in_organization(db: Session, guardian_ids: list[UUID], organization_id: UUID) -> list[Guardian]:
    if not guardian_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Debes asignar al menos un tutor.")
    unique_ids = list(dict.fromkeys(guardian_ids))
    if len(unique_ids) != len(guardian_ids):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se puede asignar el mismo tutor dos veces.")
    guardians = db.scalars(
        select(Guardian).where(Guardian.organization_id == organization_id, Guardian.id.in_(unique_ids))
    ).all()
    if len(guardians) != len(unique_ids):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uno o más tutores no pertenecen a tu centro educativo.")
    return guardians


def _student_code_duplicate_exists(
    db: Session,
    organization_id: UUID,
    student_code: str | None,
    exclude_id: UUID | None = None,
) -> bool:
    if not student_code:
        return False
    query = select(Student.id).where(
        Student.organization_id == organization_id,
        Student.student_code == student_code,
    )
    if exclude_id:
        query = query.where(Student.id != exclude_id)
    return db.scalar(query.limit(1)) is not None


def _cell_text(value) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _cell_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    text = (_cell_text(value) or "si").lower()
    return text in {"si", "sí", "s", "yes", "y", "true", "1", "activo", "activa"}


def _find_course_by_excel_values(
    db: Session,
    organization_id: UUID,
    course_name: str,
    section: str | None,
    academic_year: str | None,
) -> Course | None:
    query = select(Course).where(
        Course.organization_id == organization_id,
        func.lower(Course.name) == course_name.lower(),
    )
    if section:
        query = query.where(Course.section == section)
    else:
        query = query.where(Course.section.is_(None))
    if academic_year:
        query = query.where(Course.academic_year == academic_year)
    else:
        query = query.where(Course.academic_year.is_(None))
    return db.scalar(query.limit(1))


def _find_primary_guardian_name(db: Session, student_id: UUID) -> tuple[str | None, str | None]:
    relation = db.scalar(
        select(StudentGuardian)
        .where(StudentGuardian.student_id == student_id, StudentGuardian.is_primary.is_(True))
        .limit(1)
    )
    if not relation:
        return None, None
    guardian = db.get(Guardian, relation.guardian_id)
    if not guardian:
        return None, None
    return guardian.full_name, guardian.phone


def _set_primary_guardian(db: Session, student: Student, guardian: Guardian) -> None:
    db.execute(update(StudentGuardian).where(StudentGuardian.student_id == student.id).values(is_primary=False))
    relation = db.scalar(
        select(StudentGuardian).where(
            StudentGuardian.student_id == student.id,
            StudentGuardian.guardian_id == guardian.id,
        )
    )
    if relation:
        relation.is_primary = True
    else:
        db.add(StudentGuardian(student_id=student.id, guardian_id=guardian.id, is_primary=True))


def _style_students_worksheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor=EXCEL_HEADER_FILL)
    header_font = Font(color=EXCEL_HEADER_FONT, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_alignment

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), EXCEL_MAX_COLUMN_WIDTH)


def _student_export_rows(db: Session, organization_id: UUID) -> list[list[str | None]]:
    rows = db.execute(
        select(Student, Course)
        .join(Course, Course.id == Student.course_id)
        .where(Student.organization_id == organization_id)
        .order_by(Student.full_name)
    ).all()
    data: list[list[str | None]] = []
    for student, course in rows:
        guardian_name, guardian_phone = _find_primary_guardian_name(db, student.id)
        data.append(
            [
                student.full_name,
                student.student_code,
                course.name,
                course.section,
                course.academic_year,
                "si" if student.is_active else "no",
                guardian_name,
                guardian_phone,
            ]
        )
    return data


def _validate_import_headers(headers: list[str | None]) -> dict[str, int]:
    missing_headers = [header for header in STUDENT_IMPORT_REQUIRED_HEADERS if header not in headers]
    if missing_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Faltan columnas obligatorias: {', '.join(missing_headers)}.",
        )
    return {header: headers.index(header) for header in headers if header}


def _import_student_rows(
    *,
    rows: list[dict[str, str | None]],
    db: Session,
    current_user: User,
    request: Request,
) -> StudentImportResponse:
    created = 0
    updated = 0
    errors: list[str] = []

    for row_number, values in enumerate(rows, start=2):
        full_name = values.get("nombre_completo")
        student_code = values.get("codigo")
        course_name = values.get("curso")
        section = values.get("seccion")
        academic_year = values.get("anio_academico")
        guardian_phone = clean_phone_number(values.get("tutor_principal_telefono") or "")

        if not any(values.values()):
            continue
        if not full_name or not course_name or not guardian_phone:
            errors.append(f"Fila {row_number}: nombre_completo, curso y tutor_principal_telefono son obligatorios.")
            continue

        course = _find_course_by_excel_values(db, current_user.organization_id, course_name, section, academic_year)
        if not course:
            errors.append(f"Fila {row_number}: el curso indicado no existe en este centro.")
            continue

        guardian = db.scalar(
            select(Guardian).where(
                Guardian.organization_id == current_user.organization_id,
                Guardian.phone == guardian_phone,
                Guardian.is_active.is_(True),
            )
        )
        if not guardian:
            errors.append(f"Fila {row_number}: el tutor principal no existe o está inactivo.")
            continue

        student = None
        if student_code:
            student = db.scalar(
                select(Student).where(
                    Student.organization_id == current_user.organization_id,
                    Student.student_code == student_code,
                )
            )

        if student:
            student.full_name = full_name
            student.course_id = course.id
            student.is_active = _cell_bool(values.get("activo"))
            updated += 1
        else:
            student = Student(
                organization_id=current_user.organization_id,
                course_id=course.id,
                full_name=full_name,
                student_code=student_code,
                is_active=_cell_bool(values.get("activo")),
            )
            db.add(student)
            db.flush()
            created += 1

        _set_primary_guardian(db, student, guardian)

    create_audit_log(
        db,
        action="import_students_excel",
        user=current_user,
        organization_id=current_user.organization_id,
        entity_name="students",
        new_data={"created": created, "updated": updated, "errors": errors},
        request=request,
    )
    db.commit()
    return StudentImportResponse(created=created, updated=updated, errors=errors)


@router.post("", response_model=StudentResponse, status_code=status.HTTP_201_CREATED)
def create_student(
    payload: StudentCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_admin(current_user)
    organization_id = current_user.organization_id
    _ensure_course_in_organization(db, payload.course_id, organization_id)
    _ensure_guardians_in_organization(db, payload.guardian_ids, organization_id)
    if _student_code_duplicate_exists(db, organization_id, payload.student_code):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Ya existe un estudiante con ese código.")
    student = Student(
        organization_id=organization_id,
        course_id=payload.course_id,
        full_name=payload.full_name,
        student_code=payload.student_code,
        is_active=payload.is_active,
    )
    db.add(student)
    db.flush()
    primary_guardian_id = payload.primary_guardian_id or (payload.guardian_ids[0] if payload.guardian_ids else None)
    db.add_all(
        [
            StudentGuardian(
                student_id=student.id,
                guardian_id=guardian_id,
                is_primary=guardian_id == primary_guardian_id,
            )
            for guardian_id in payload.guardian_ids
        ]
    )
    create_audit_log(
        db,
        action="create_student",
        user=current_user,
        organization_id=organization_id,
        entity_name="students",
        entity_id=student.id,
        new_data={"full_name": student.full_name, "course_id": str(student.course_id)},
        request=request,
    )
    db.commit()
    db.refresh(student)
    return student


@router.get("", response_model=list[StudentResponse])
def list_students(
    search: str | None = None,
    course_id: UUID | None = None,
    is_active: bool | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_user(current_user)
    query = select(Student).where(Student.organization_id == current_user.organization_id)
    if search:
        query = query.where(Student.full_name.ilike(f"%{search}%"))
    if course_id:
        query = query.where(Student.course_id == course_id)
    if is_active is not None:
        query = query.where(Student.is_active == is_active)
    return db.scalars(query.order_by(Student.full_name)).all()


@router.get("/export")
def export_students_excel(
    file_format: str = Query(default="xlsx", pattern="^(xlsx|csv)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_user(current_user)
    export_rows = _student_export_rows(db, current_user.organization_id)
    if file_format == "csv":
        text_stream = BytesIO()
        string_stream = StringIO()
        writer = csv.writer(string_stream, lineterminator="\r\n")
        writer.writerow(STUDENT_EXCEL_HEADERS)
        writer.writerows(export_rows)
        text_stream.write((CSV_BOM + string_stream.getvalue()).encode("utf-8"))
        text_stream.seek(0)
        return StreamingResponse(
            text_stream,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="estudiantes.csv"'},
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Estudiantes"
    worksheet.append(STUDENT_EXCEL_HEADERS)

    for row in export_rows:
        worksheet.append(row)

    _style_students_worksheet(worksheet)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="estudiantes.xlsx"'},
    )


@router.post("/import", response_model=StudentImportResponse)
def import_students_excel(
    request: Request,
    file: UploadFile,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_admin(current_user)
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Debes subir un archivo .xlsx o .csv.")
    filename = file.filename.lower()
    content = file.file.read()
    rows: list[dict[str, str | None]]

    if filename.endswith(".xlsx"):
        try:
            workbook = load_workbook(BytesIO(content), data_only=True)
        except (InvalidFileException, OSError, ValueError):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo Excel no es válido.")

        worksheet = workbook.active
        headers = [_cell_text(cell.value) for cell in worksheet[1]]
        header_index = _validate_import_headers(headers)
        rows = [
            {
                header: _cell_text(row[index]) if index < len(row) else None
                for header, index in header_index.items()
            }
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
    elif filename.endswith(".csv"):
        try:
            csv_text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo CSV debe estar en UTF-8.")

        reader = csv.DictReader(StringIO(csv_text))
        headers = [_cell_text(header) for header in (reader.fieldnames or [])]
        _validate_import_headers(headers)
        rows = [
            {header: _cell_text(row.get(header)) for header in headers if header}
            for row in reader
        ]
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Debes subir un archivo .xlsx o .csv.")

    return _import_student_rows(rows=rows, db=db, current_user=current_user, request=request)


@router.get("/{student_id}", response_model=StudentResponse)
def get_student(
    student_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_user(current_user)
    return _get_student_or_404(db, student_id, current_user.organization_id)


@router.put("/{student_id}", response_model=StudentResponse)
def update_student(
    student_id: UUID,
    payload: StudentUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_admin(current_user)
    student = _get_student_or_404(db, student_id, current_user.organization_id)
    old_data = {
        "course_id": str(student.course_id),
        "full_name": student.full_name,
        "student_code": student.student_code,
        "is_active": student.is_active,
    }
    update_data = payload.model_dump(exclude_unset=True)
    if "course_id" in update_data:
        _ensure_course_in_organization(db, update_data["course_id"], current_user.organization_id)
    if "student_code" in update_data and _student_code_duplicate_exists(
        db, current_user.organization_id, update_data["student_code"], exclude_id=student.id
    ):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Ya existe un estudiante con ese código.")
    for field, value in update_data.items():
        setattr(student, field, value)
    create_audit_log(
        db,
        action="update_student",
        user=current_user,
        organization_id=current_user.organization_id,
        entity_name="students",
        entity_id=student.id,
        old_data=old_data,
        new_data={key: str(value) if isinstance(value, UUID) else value for key, value in update_data.items()},
        request=request,
    )
    db.commit()
    db.refresh(student)
    return student


@router.delete("/{student_id}", response_model=StudentResponse)
def delete_student(
    student_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_admin(current_user)
    student = _get_student_or_404(db, student_id, current_user.organization_id)
    student.is_active = False
    create_audit_log(
        db,
        action="delete_student",
        user=current_user,
        organization_id=current_user.organization_id,
        entity_name="students",
        entity_id=student.id,
        old_data={"is_active": True},
        new_data={"is_active": False},
        request=request,
    )
    db.commit()
    db.refresh(student)
    return student


@router.post("/{student_id}/guardians", response_model=StudentGuardianResponse, status_code=status.HTTP_201_CREATED)
def assign_guardian(
    student_id: UUID,
    payload: StudentGuardianCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_admin(current_user)
    student = _get_student_or_404(db, student_id, current_user.organization_id)
    _ensure_guardians_in_organization(db, [payload.guardian_id], current_user.organization_id)
    existing = db.scalar(
        select(StudentGuardian).where(
            StudentGuardian.student_id == student.id,
            StudentGuardian.guardian_id == payload.guardian_id,
        )
    )
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="El tutor ya está asignado al estudiante.")
    if payload.is_primary:
        db.execute(update(StudentGuardian).where(StudentGuardian.student_id == student.id).values(is_primary=False))
    relation = StudentGuardian(
        student_id=student.id,
        guardian_id=payload.guardian_id,
        is_primary=payload.is_primary,
    )
    db.add(relation)
    db.commit()
    db.refresh(relation)
    return relation


@router.get("/{student_id}/guardians", response_model=list[StudentGuardianResponse])
def list_student_guardians(
    student_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_user(current_user)
    student = _get_student_or_404(db, student_id, current_user.organization_id)
    return db.scalars(select(StudentGuardian).where(StudentGuardian.student_id == student.id)).all()


@router.delete("/{student_id}/guardians/{guardian_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_student_guardian(
    student_id: UUID,
    guardian_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_admin(current_user)
    student = _get_student_or_404(db, student_id, current_user.organization_id)
    relation = db.scalar(
        select(StudentGuardian).where(
            StudentGuardian.student_id == student.id,
            StudentGuardian.guardian_id == guardian_id,
        )
    )
    if not relation:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Relación estudiante-tutor no encontrada.")
    db.delete(relation)
    db.commit()


@router.put("/{student_id}/guardians/{guardian_id}/primary", response_model=StudentGuardianResponse)
def set_primary_guardian(
    student_id: UUID,
    guardian_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_school_admin(current_user)
    student = _get_student_or_404(db, student_id, current_user.organization_id)
    relation = db.scalar(
        select(StudentGuardian).where(
            StudentGuardian.student_id == student.id,
            StudentGuardian.guardian_id == guardian_id,
        )
    )
    if not relation:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Relación estudiante-tutor no encontrada.")
    db.execute(update(StudentGuardian).where(StudentGuardian.student_id == student.id).values(is_primary=False))
    relation.is_primary = True
    db.commit()
    db.refresh(relation)
    return relation
